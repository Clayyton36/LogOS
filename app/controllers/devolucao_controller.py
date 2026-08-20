from datetime import datetime

import openpyxl
from openpyxl.utils import get_column_letter

from app.models.devolucao import Devolucao
from app.repositories.devolucao_repository import DevolucaoRepository


class DevolucaoValidationError(Exception):
    pass


DESTINOS_VALIDOS = ("ESTOQUE", "TROCA", "DESCARTE")

PLATAFORMAS_VALIDAS = ("MERCADO LIVRE", "SHOPEE", "TIKTOK", "CORREIOS", "FISICO")

OBSERVACOES_RECEBIMENTO_VALIDAS = ("COM AVARIA", "SEM AVARIA", "PRODUTO TROCADO", "ENVIADO ERRADO")

ACESSORIOS_VALIDOS = ("OK", "DANIFICADO")

AVARIA_VALIDOS = ("AVARIADO", "SEM AVARIA")

LANCADO_SISTEMA_VALIDOS = ("SIM", "NAO")

INDICADOR_FILTROS = {
    "recebidas": {"status": "Recebida"},
    "aguardando_decisao": {"status": "Analisada"},
    "estoque": {"status": "Finalizada", "destino": "ESTOQUE"},
    "troca": {"status": "Finalizada", "destino": "TROCA"},
    "descarte": {"status": "Finalizada", "destino": "DESCARTE"},
}

COLUNAS_EXPORTACAO = [
    ("numero_pedido", "Número do pedido"),
    ("numero_nf", "Número da NF"),
    ("cliente", "Cliente"),
    ("plataforma", "Plataforma"),
    ("sku", "SKU"),
    ("produto", "Produto"),
    ("responsavel_recebimento", "Responsável pelo recebimento"),
    ("data_recebimento", "Data do recebimento"),
    ("condicao_produto", "Condição do produto"),
    ("avaria", "Avaria"),
    ("acessorios", "Acessórios"),
    ("situacao_encontrada", "Situação encontrada"),
    ("observacoes_analise", "Observações da análise"),
    ("data_analise", "Data da análise"),
    ("destino", "Destino"),
    ("observacoes_decisao", "Observações da decisão"),
    ("data_decisao", "Data da decisão"),
    ("status", "Status"),
    ("observacoes", "Observações do recebimento"),
    ("lancado_sistema", "Lançado no sistema"),
]


class DevolucaoController:

    def __init__(self, repository: DevolucaoRepository = None):
        self.repository = repository or DevolucaoRepository()

    def registrar_recebimento(
        self,
        numero_pedido: str,
        plataforma: str,
        cliente: str,
        produto: str,
        responsavel_recebimento: str,
        numero_nf: str = "",
        observacoes: str = "",
    ) -> int:
        campos_obrigatorios = {
            "Número do pedido": numero_pedido,
            "Plataforma": plataforma,
            "Cliente": cliente,
            "Responsável pelo recebimento": responsavel_recebimento,
        }
        faltando = [nome for nome, valor in campos_obrigatorios.items() if not valor.strip()]
        if faltando:
            raise DevolucaoValidationError(
                "Preencha os campos obrigatórios: " + ", ".join(faltando)
            )

        if plataforma.strip().upper() not in PLATAFORMAS_VALIDAS:
            raise DevolucaoValidationError(
                "Selecione uma plataforma válida: " + ", ".join(PLATAFORMAS_VALIDAS)
            )

        devolucao = Devolucao(
            numero_pedido=numero_pedido.strip().upper(),
            plataforma=plataforma.strip().upper(),
            cliente=cliente.strip().upper(),
            produto=produto.strip().upper(),
            responsavel_recebimento=responsavel_recebimento.strip().upper(),
            numero_nf=numero_nf.strip().upper(),
            observacoes=observacoes.strip().upper(),
            status="Recebida",
            data_recebimento=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        )

        return self.repository.salvar(devolucao)

    def consultar(self, termo_busca: str = "", plataforma: str = "", lancado_sistema: str = ""):
        return self.repository.listar(
            termo_busca=termo_busca.strip(),
            plataforma=plataforma.strip(),
            lancado_sistema=lancado_sistema.strip(),
        )

    def marcar_lancado_sistema(self, devolucao_id: int, lancado_sistema: str):
        if devolucao_id is None:
            raise DevolucaoValidationError("Selecione uma devolução.")
        if lancado_sistema not in LANCADO_SISTEMA_VALIDOS:
            raise DevolucaoValidationError("Valor inválido para lançado no sistema.")
        self.repository.atualizar_lancado_sistema(devolucao_id, lancado_sistema)

    def listar_plataformas(self):
        return self.repository.listar_plataformas()

    def obter_indicadores(self):
        return self.repository.obter_indicadores()

    def listar_por_indicador(self, chave: str):
        filtro = INDICADOR_FILTROS.get(chave, {})
        return self.repository.listar_por_indicador(
            status=filtro.get("status"), destino=filtro.get("destino")
        )

    def reabrir_devolucao(self, devolucao_id: int):
        if devolucao_id is None:
            raise DevolucaoValidationError("Selecione uma devolução para reabrir.")
        self.repository.reabrir_decisao(devolucao_id)

    def exportar_para_excel(
        self, caminho: str, termo_busca: str = "", plataforma: str = "", lancado_sistema: str = ""
    ) -> int:
        devolucoes = self.repository.listar(
            termo_busca=termo_busca.strip(),
            plataforma=plataforma.strip(),
            lancado_sistema=lancado_sistema.strip(),
        )

        planilha = openpyxl.Workbook()
        aba = planilha.active
        aba.title = "Devoluções"

        aba.append([titulo for _, titulo in COLUNAS_EXPORTACAO])
        for devolucao in devolucoes:
            aba.append([getattr(devolucao, campo) for campo, _ in COLUNAS_EXPORTACAO])

        for indice, (_, titulo) in enumerate(COLUNAS_EXPORTACAO, start=1):
            aba.column_dimensions[get_column_letter(indice)].width = max(len(titulo) + 2, 14)

        planilha.save(caminho)
        return len(devolucoes)

    def listar_pendentes_analise(self):
        return self.repository.listar_por_status("Recebida")

    def registrar_analise(
        self,
        devolucao_id: int,
        condicao_produto: str,
        avaria: str = "",
        acessorios: str = "",
        situacao_encontrada: str = "",
    ):
        if devolucao_id is None:
            raise DevolucaoValidationError("Selecione uma devolução para analisar.")

        if not condicao_produto.strip():
            raise DevolucaoValidationError("Preencha o campo obrigatório: Condição do produto.")

        self.repository.atualizar_analise(
            devolucao_id=devolucao_id,
            condicao_produto=condicao_produto.strip(),
            avaria=avaria.strip(),
            acessorios=acessorios.strip(),
            situacao_encontrada=situacao_encontrada.strip(),
            observacoes_analise="",
            data_analise=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            novo_status="Analisada",
        )

    def listar_pendentes_decisao(self):
        return self.repository.listar_por_status("Analisada")

    def registrar_decisao(
        self,
        devolucao_id: int,
        destino: str,
        observacoes_decisao: str = "",
    ):
        if devolucao_id is None:
            raise DevolucaoValidationError("Selecione uma devolução para decidir o destino.")

        if destino not in DESTINOS_VALIDOS:
            raise DevolucaoValidationError(
                "Selecione um destino válido: " + ", ".join(DESTINOS_VALIDOS)
            )

        self.repository.atualizar_decisao(
            devolucao_id=devolucao_id,
            destino=destino,
            observacoes_decisao=observacoes_decisao.strip(),
            data_decisao=datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            novo_status="Finalizada",
        )
